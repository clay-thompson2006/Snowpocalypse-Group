import openpyxl
from bs4 import BeautifulSoup
import requests
import openpyxl
import tkinter as tk
import time
import datetime
import Gmail
import SMS
while True:
    current_time = datetime.datetime.now
    if current_time.hour == 2 and current_time.minute == 30:
        url = "https://weather.com/weather/tenday/l/Canfield+OH+44406"
        response = requests.get(url)
        soup = BeautifulSoup(response.text, 'html.parser')
        temp_elements = soup.find_all('span', class_ = 'DetailsSummary--lowTempValue--2tesQ',)
        dates = soup.find_all('h3', class_ = 'DetailsSummary--daypartName--kbngc',)
        temperatures = [element.get_text() for element in temp_elements]
        text_dates = [element.get_text() for element in dates]

        try:
            workbook = openpyxlload_workbook('snow_excel.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet['A1'] = 'Dates'
            sheet['B1'] = 'temperature'
        next_row = sheet.max_row + 2

        for row, (item1, item2) in enumerate(zip(text_dates, temperatures), start= next_row):
            sheet[f'A{row}'] = item1
            sheet[f'B{row}'] = item2
        workbook.save('snow_excel.xlsx')
        workbook = openpyxl.load_workbook('snow_excel.xlsx') 
        sheet = workbook['Sheet']  
        excel_data = []

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            row_data = []
            for cell in row:
                row_data.append(cell.value) 
            excel_data.append(row_data)
        workbook.close()


        root = tk.Tk()
        root.geometry("700x350")
        scroll_bar = tk.Scrollbar(root, orient='vertical')
        scroll_bar.pack(side= tk.RIGHT, fill='y')
        my_list = tk.Listbox(root)
        title = root.title("Weather thingy")
        my_list.config(yscrollcommand=scroll_bar.set)
        scroll_bar.config(command=my_list.yview)
        for item in excel_data:
            my_list.insert(tk.END, item)
        my_list.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll_bar.pack(side=tk.RIGHT, fill=tk.Y)
        root.mainloop()

if temperatures <= 10:
    Gmail.Alert()
    SMS.Alert()
